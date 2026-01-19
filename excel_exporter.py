import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 로직 파일에서 필요한 상수들을 가져옵니다
from scheduler_core import STAFF_NAMES, NIGHT_SHIFTS

def save_to_excel(schedule, cycle_starts, filename="shift_schedule.xlsx"):
    """
    schedule: 최적화된 근무표 객체
    cycle_starts: {직원인덱스: 시작오프셋} 정보 (파트너 스케줄 계산용)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "근무표"

    # --- 1. 색상 정의 (Hex Codes) ---
    COLOR_TEXT = {
        '주': '008000', # 초록
        '야': '0000FF', # 파랑
        '당': 'A52A2A', # 갈색
        '비': 'FF0000', # 빨강
        '출': '00BFFF', # 하늘색
        '휴': '000000', # 검정
        '생': 'CD853F'  # 밝은 갈색
    }

    # 배경색 (셀)
    BG_1_3_TEAM = 'FFEFD5' # 옅은 주황
    BG_2_4_TEAM = 'E0FFFF' # 옅은 파랑
    BG_CHE_SUP  = 'E6E6FA' # 옅은 보라
    BG_BO_SUP   = 'F0FFF0' # 옅은 초록
    BG_VACATION = 'FFC0CB' # 분홍색 (휴가)
    BG_PARTNER  = 'F5F5F5' # 파트너 스케줄 배경 (옅은 회색)
    BG_HEADER   = 'DDDDDD' # 헤더 배경 (회색)

    # [통계용 배경색]
    BG_STAT_VAC   = 'FFE4E1' # 휴가수 (MistyRose)
    BG_STAT_DAY   = 'F0FFF0' # 주간수 (Honeydew)
    BG_STAT_NIGHT = 'E6F2FF' # 야간수 (AliceBlue보다 진함)
    BG_STAT_TOTAL = 'FFF2CC' # 총시간 (연한 노랑/오렌지)

    # --- 2. 스타일 헬퍼 함수 ---
    def get_font(shift_char):
        color = COLOR_TEXT.get(shift_char, '000000')
        return Font(bold=True, color=color)

    def get_fill(staff_name, shift_char):
        if shift_char == '휴':
            return PatternFill(start_color=BG_VACATION, end_color=BG_VACATION, fill_type='solid')

        color_code = 'FFFFFF'
        if staff_name.startswith("["): 
             color_code = BG_PARTNER
        elif staff_name.startswith('1팀') or staff_name.startswith('3팀'):
            color_code = BG_1_3_TEAM
        elif staff_name.startswith('2팀') or staff_name.startswith('4팀'):
            color_code = BG_2_4_TEAM
        elif staff_name.startswith('체지원'):
            color_code = BG_CHE_SUP
        elif staff_name.startswith('보지원'):
            color_code = BG_BO_SUP
        
        return PatternFill(start_color=color_code, end_color=color_code, fill_type='solid')

    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # 날짜 헤더 그리기 함수
    def write_date_header(row_idx, title="구분", with_stats=False):
        # 1. 첫열 (타이틀)
        ws.cell(row=row_idx, column=1, value=title)
        
        # 2. 날짜열 (1 ~ num_days)
        for d in range(schedule.num_days):
            cell = ws.cell(row=row_idx, column=d+2, value=d+1)
            cell.alignment = center_align
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.fill = PatternFill(start_color=BG_HEADER, end_color=BG_HEADER, fill_type='solid')
        
        # 3. 통계 헤더 (with_stats=True일 때만 표시)
        if with_stats:
            stat_headers = [("휴가", BG_STAT_VAC), ("주간", BG_STAT_DAY), 
                            ("야간", BG_STAT_NIGHT), ("총시간", BG_STAT_TOTAL)]
            
            # [변경점] 한 칸 건너뛰고 시작 (공백 열 추가)
            start_col = schedule.num_days + 3 
            
            for i, (text, bg_color) in enumerate(stat_headers):
                cell = ws.cell(row=row_idx, column=start_col + i, value=text)
                cell.alignment = center_align
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

    # --- 3. 엑셀 작성 ---
    
    current_row = 1
    
    # [상단 헤더] 파트너용
    write_date_header(current_row, "기준표", with_stats=False)
    current_row += 1

    # --- [섹션 1] 파트너(기준) 스케줄 ---
    PARTNER_TEAMS = [
        ("[1팀]", "1팀체계"),
        ("[2팀]", "2팀체계"),
        ("[3팀]", "3팀체계"),
        ("[4팀]", "4팀체계")
    ]
    VISUAL_CYCLE = ['주', '야', '비', '생']
    
    for display_name, rep_name in PARTNER_TEAMS:
        name_cell = ws.cell(row=current_row, column=1, value=display_name)
        name_cell.alignment = center_align
        name_cell.border = thin_border
        name_cell.font = Font(bold=True, italic=True)
        name_cell.fill = get_fill(display_name, '')

        if rep_name in STAFF_NAMES:
            rep_idx = STAFF_NAMES.index(rep_name)
            start_offset = cycle_starts.get(rep_idx, 0)
        else:
            start_offset = 0

        for c in range(schedule.num_days):
            col_idx = c + 2
            cycle_idx = (c + start_offset) % 4
            shift_val = VISUAL_CYCLE[cycle_idx]
            
            cell = ws.cell(row=current_row, column=col_idx, value=shift_val)
            cell.font = get_font(shift_val)
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = get_fill(display_name, shift_val)
            
        current_row += 1

    # --- [공백 행] ---
    current_row += 1 
    
    # [하단 헤더] 실제 근무표용 (통계 포함)
    write_date_header(current_row, "직원명", with_stats=True)
    current_row += 1

    # --- [섹션 2] 실제 근무표 + 통계 계산 ---
    for r, name in enumerate(STAFF_NAMES):
        # 1. 이름
        name_cell = ws.cell(row=current_row, column=1, value=name)
        name_cell.alignment = center_align
        name_cell.border = thin_border
        name_cell.font = Font(bold=True)
        name_cell.fill = get_fill(name, '')

        # 통계 카운터
        cnt_vacation = 0 
        cnt_day = 0      
        cnt_night = 0    

        # 2. 날짜별 데이터
        for c in range(schedule.num_days):
            col_idx = c + 2
            raw_val = schedule.grid[r][c]
            
            if raw_val == '휴': cnt_vacation += 1
            elif raw_val in ['출', '주']: cnt_day += 1
            elif raw_val in ['당', '야']: cnt_night += 1
            
            display_val = raw_val
            if c > 0 and schedule.grid[r][c-1] in NIGHT_SHIFTS and raw_val == '생':
                display_val = '비'
            
            cell = ws.cell(row=current_row, column=col_idx, value=display_val)
            cell.font = get_font(display_val)
            cell.fill = get_fill(name, display_val)
            cell.alignment = center_align
            cell.border = thin_border
        
        # 3. 통계 출력
        total_hours = (cnt_day * 8) + (cnt_night * 13)
        stat_values = [cnt_vacation, cnt_day, cnt_night, total_hours]
        stat_bgs = [BG_STAT_VAC, BG_STAT_DAY, BG_STAT_NIGHT, BG_STAT_TOTAL]
        
        # [변경점] 한 칸 건너뛰고 시작
        start_stat_col = schedule.num_days + 3
        
        for i, val in enumerate(stat_values):
            cell = ws.cell(row=current_row, column=start_stat_col + i, value=val)
            cell.alignment = center_align
            cell.border = thin_border
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=stat_bgs[i], end_color=stat_bgs[i], fill_type='solid')
            
        current_row += 1

    # 열 너비 자동 조정
    ws.column_dimensions['A'].width = 15
    
    # 날짜열 너비
    for d in range(schedule.num_days):
        col_letter = get_column_letter(d+2)
        ws.column_dimensions[col_letter].width = 4
    
    # [변경점] 공백열(Separator) 너비 좁게 설정 (2)
    sep_col_letter = get_column_letter(schedule.num_days + 2)
    ws.column_dimensions[sep_col_letter].width = 2
    
    # 통계열 너비
    for i in range(4):
        col_letter = get_column_letter(schedule.num_days + 3 + i)
        ws.column_dimensions[col_letter].width = 8

    # 저장
    try:
        wb.save(filename)
        print(f"\n[성공] 근무표가 '{filename}' 파일로 저장되었습니다!")
        print(" -> 특징: 근무표와 통계 사이 공백열 추가")
    except PermissionError:
        print(f"\n[오류] '{filename}' 파일이 열려있습니다. 닫고 다시 실행해주세요.")